from openpyxl import load_workbook

# Replace 'your_excel_file.xlsx' and 'new_excel_file.xlsx' with your actual file names
excel_file_path = 'your_excel_file.xlsx'
new_excel_file_path = 'new_excel_file.xlsx'

# Load the workbook
workbook = load_workbook(excel_file_path)

# Create a new workbook to save the changes
new_workbook = load_workbook()

# Copy all sheets to the new workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    new_workbook.create_sheet(title=sheet_name)
    new_sheet = new_workbook[sheet_name]

    # Copy all cells from the original sheet to the new sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value

# Save the new workbook with a new name
new_workbook.save(new_excel_file_path)

print(f'Excel file saved as: {new_excel_file_path}')



import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('your_excel_file.xlsx')

# Iterate through all sheets in the workbook
for sheet in workbook.sheetnames:
    # Get the active sheet
    current_sheet = workbook[sheet]

    # Recalculate all formulas in the sheet
    for row in current_sheet:
        for cell in row:
            if cell.formula:
                cell.value = cell.calculate_value()

# Save the workbook
workbook.save('your_updated_excel_file.xlsx')
